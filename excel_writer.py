"""Genera el Excel final con 3 hojas y formato profesional."""

from pathlib import Path
import pandas as pd
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_ERROR_FILL   = PatternFill("solid", fgColor="FF0000")
_WARN_FILL    = PatternFill("solid", fgColor="FFC000")
_OK_FILL      = PatternFill("solid", fgColor="70AD47")
_WHITE_FONT   = Font(color="FFFFFF", bold=True)
_THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_MONEY_FMT    = '#,##0.00'
_PCT_FMT      = '0.00"%"'


def _style_header(ws):
    for cell in ws[1]:
        cell.font      = _WHITE_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER


def _autofit(ws, max_width=45):
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, max_width)


def _apply_money_cols(ws, col_names: list[str], headers: list[str]):
    for idx, h in enumerate(headers, start=1):
        if h in col_names:
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in row:
                    cell.number_format = _MONEY_FMT


def write_excel(
    df_base: pd.DataFrame,
    df_val: pd.DataFrame,
    df_prorrateo: pd.DataFrame,
    output_path: Path,
):
    """Escribe las 3 hojas en output_path."""
    money_base = ["subtotal", "iva_19", "iva_5", "total"]
    money_val  = ["total"]
    money_pror = [
        "iva_total", "iva_mandatos", "iva_base_prorateo",
        "ingresos_gravados", "ingresos_excluidos",
        "iva_descontable", "iva_no_descontable",
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Hoja 1: BASE_DATOS ──────────────────────────────────────────
        df_base.to_excel(writer, sheet_name="BASE_DATOS", index=False)
        ws1 = writer.sheets["BASE_DATOS"]
        _style_header(ws1)
        _apply_money_cols(ws1, money_base, list(df_base.columns))
        _autofit(ws1)
        ws1.freeze_panes = "A2"

        # ── Hoja 2: VALIDACION ──────────────────────────────────────────
        df_val.to_excel(writer, sheet_name="VALIDACION", index=False)
        ws2 = writer.sheets["VALIDACION"]
        _style_header(ws2)
        _apply_money_cols(ws2, money_val, list(df_val.columns))

        val_col = list(df_val.columns).index("validacion") + 1 if "validacion" in df_val.columns else None
        if val_col:
            for row in ws2.iter_rows(min_row=2, min_col=val_col, max_col=val_col):
                for cell in row:
                    if str(cell.value).upper() == "ERROR":
                        cell.fill = _ERROR_FILL
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif str(cell.value).upper() == "OK":
                        cell.fill = _OK_FILL

        _autofit(ws2)
        ws2.freeze_panes = "A2"

        # ── Hoja 3: PRORRATEO_IVA ───────────────────────────────────────
        df_prorrateo.to_excel(writer, sheet_name="PRORRATEO_IVA", index=False)
        ws3 = writer.sheets["PRORRATEO_IVA"]
        _style_header(ws3)
        _apply_money_cols(ws3, money_pror, list(df_prorrateo.columns))

        pct_col_idx = (
            list(df_prorrateo.columns).index("pct_prorateo") + 1
            if "pct_prorateo" in df_prorrateo.columns else None
        )
        if pct_col_idx:
            for row in ws3.iter_rows(min_row=2, min_col=pct_col_idx, max_col=pct_col_idx):
                for cell in row:
                    cell.number_format = _PCT_FMT

        _autofit(ws3)
        ws3.freeze_panes = "A2"

    return output_path
