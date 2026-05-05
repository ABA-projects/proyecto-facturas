"""Genera Excel Formato 1003 DIAN desde DataFrame de exógenas."""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_ICA_FILL    = PatternFill("solid", fgColor="FFF2CC")   # amarillo claro para ICA excluido
_WHITE_FONT  = Font(color="FFFFFF", bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_MONEY_FMT = "#,##0"
_PCT_FMT   = '0.00"%"'

_COLS_1003 = [
    ("concepto",        "Concepto"),
    ("tipo_doc",        "Tipo\nDocumento"),
    ("nit",             "Número\nIdentificación"),
    ("dv",              "DV"),
    ("primer_apellido", "Primer\nApellido"),
    ("segundo_apellido","Segundo\nApellido"),
    ("primer_nombre",   "Primer\nNombre"),
    ("otros_nombres",   "Otros\nNombres"),
    ("razon_social",    "Razón Social"),
    ("direccion",       "Dirección"),
    ("cod_dpto",        "Cód.\nDpto"),
    ("cod_mpio",        "Cód.\nMpio"),
    ("base",            "Base Sujeta\na Retención"),
    ("retencion",       "Retención\nPracticada"),
    ("porcentaje",      "Porcentaje\n%"),
]

_COLS_DETALLE = [
    "archivo", "tipo_cert", "razon_social", "nit", "dv", "tipo_doc",
    "direccion", "ciudad_retencion", "concepto", "base", "retencion", "porcentaje", "error",
]


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font      = _WHITE_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER
    ws.row_dimensions[1].height = 32


def _autofit(ws, max_width: int = 40) -> None:
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, max_width)


def write_1003(df_1003: pd.DataFrame, df_detalle: pd.DataFrame, output_path: Path) -> Path:
    """
    Escribe Excel con dos hojas:
      FORMATO_1003 — filas agregadas por (nit, concepto)
      DETALLE      — un registro por PDF procesado
    """
    wb = Workbook()

    # ── Hoja FORMATO_1003 ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "FORMATO_1003"

    field_names = [f for f, _ in _COLS_1003]
    headers     = [h for _, h in _COLS_1003]
    ws.append(headers)

    for _, row in df_1003.iterrows():
        ws.append([row.get(f, "") for f in field_names])

    _style_header(ws)
    _autofit(ws)

    # Formato monetario en Base y Retención
    base_col = field_names.index("base") + 1
    ret_col  = field_names.index("retencion") + 1
    pct_col  = field_names.index("porcentaje") + 1

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in r:
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
        ws.cell(r[0].row, base_col).number_format = _MONEY_FMT
        ws.cell(r[0].row, ret_col).number_format  = _MONEY_FMT
        ws.cell(r[0].row, pct_col).number_format  = _PCT_FMT

    ws.freeze_panes = "A2"

    # Fila de totales
    last = ws.max_row + 1
    ws.cell(last, base_col, f"=SUM({get_column_letter(base_col)}2:{get_column_letter(base_col)}{ws.max_row})")
    ws.cell(last, ret_col,  f"=SUM({get_column_letter(ret_col)}2:{get_column_letter(ret_col)}{ws.max_row})")
    ws.cell(last, 1, "TOTAL")
    for c in [1, base_col, ret_col]:
        cell = ws.cell(last, c)
        cell.font   = Font(bold=True)
        cell.fill   = PatternFill("solid", fgColor="DCE6F1")
        cell.border = _THIN_BORDER
    ws.cell(last, base_col).number_format = _MONEY_FMT
    ws.cell(last, ret_col).number_format  = _MONEY_FMT

    # ── Hoja DETALLE ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("DETALLE")
    det_headers = [c.replace("_", " ").title() for c in _COLS_DETALLE]
    ws2.append(det_headers)

    for _, row in df_detalle.iterrows():
        ws2.append([row.get(c, "") for c in _COLS_DETALLE])

    _style_header(ws2)
    _autofit(ws2, max_width=50)

    # Resaltar filas con error o ICA en DETALLE
    for r in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        tipo_idx = _COLS_DETALLE.index("tipo_cert")
        err_idx  = _COLS_DETALLE.index("error")
        tipo_val = str(r[tipo_idx].value or "")
        err_val  = str(r[err_idx].value  or "")
        fill = _ICA_FILL if tipo_val == "ICA" else None
        for cell in r:
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws2.freeze_panes = "A2"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
